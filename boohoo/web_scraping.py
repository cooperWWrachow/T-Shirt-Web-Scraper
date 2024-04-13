import re
from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl.styles import Alignment, Font
import os


user_input  = input("input size, max price, and a color (optional) with a space in between each: ")
filters = user_input.split()

if len(filters) == 3:
    size = filters[0].upper()
    max_price = int(filters[1])
    color = filters[2].capitalize() # url only accepst colors with first letter capitalized
elif len(filters) == 2:
    size = filters[0].upper()
    max_price = int(filters[1])
else:
    raise ValueError("Input must contain at least a size and max price.")
    

url = f"https://www.boohooman.com/us/search?q=oversized%20tshirt&prefn1=classification&prefv1=boohooMAN%20Tall%7CMain%20Collection&prefn2=sizeRefinement&prefv2={size}&sz=80"

page = requests.get(url).text
doc = BeautifulSoup(page, "html.parser")

total_pages = int(doc.find(class_="pagination-info js-pagination-info hidden-on-mobile hidden-on-tablet-portrait").string.strip()[10:])
per_page = 80

items_found = {}

# iterate through each page based on total total items per page and page count
for page in range(total_pages):
    start = page * per_page
    # if user requests a color, adjust url accordingly
    if len(filters) == 3:
        url = f"https://www.boohooman.com/us/search?q=oversized%20tshirt&prefn1=classification&prefv1=boohooMAN%20Tall%7CMain%20Collection&prefn2=color&prefv2={color}&prefn3=sizeRefinement&prefv3={size}&start={start}&sz={per_page}"
    else:
        url = f"https://www.boohooman.com/us/search?q=oversized%20tshirt&prefn1=classification&prefv1=boohooMAN%20Tall%7CMain%20Collection&prefn2=sizeRefinement&prefv2={size}&start={start}&sz={per_page}"
    
    page = requests.get(url).text
    doc = BeautifulSoup(page, "html.parser")

    div = doc.find(class_="search-result-content js-search-result-content")

    # all instances where there is a string within a tag that states "oversized"
    items = div.find_all(string = re.compile("Oversized"))
    for item in items:
        # correct URL is several parents up the tree
        parents = item.parent.parent.parent
        link_class = parents.find(class_="product-image js-product-image load-bg").a
        link = "https://www.boohooman.com" + link_class['data-href']
        
        next_parent = item.find_parent(class_="grid-tile")

        # handles products NOT on sale that throw a None type error (dont have a sales price)
        try:
            sale_price = next_parent.find(class_="product-sales-price product-sales-price-percent")
            price = int(sale_price.contents[0].strip().replace("$", "")[:-3])
            if price <= max_price:
                items_found[item.replace("\n", "")] = {"price": price, "link": link}
            else:
                continue
        except:
            pass


# place into excel

workbook = openpyxl.load_workbook('boohoo.xlsx')

std = workbook['Sheet1']

# clears entire spreadsheet
for i in std.iter_rows():
    for cell in i:
        cell.value = None
        cell.hyperlink = None 


# populate headers
col_titles = ['Name', 'Price', 'URL']
for column_index, title in enumerate(col_titles, start=1):
    header = std.cell(row=1, column=column_index, value=title)
    header.alignment = Alignment(horizontal='center', vertical='center')


count = 0
# populate with data 
for row_index, (product, details) in enumerate(items_found.items(), start=3):
    name_cell = std.cell(row=row_index, column=1, value=product) 
    price_cell = std.cell(row=row_index, column=2, value=details['price'])  
    url_cell = std.cell(row=row_index, column=3, value=details['link'])  

    # adjust width of columns A and C for name and link
    product_width = len(product) + 2  
    std.column_dimensions['A'].width = max(std.column_dimensions['A'].width, product_width)

    price_cell.alignment = Alignment(horizontal='center', vertical='center')

    link_width = len(details['link']) + 2 
    std.column_dimensions['C'].width = max(std.column_dimensions['C'].width, link_width)
    url_cell.hyperlink = details['link']
    url_cell.font = Font(color='0000ff', underline='single')
    count += 1

# set the count variable then save 
count_cell = std['D1']
count_cell.alignment = Alignment(horizontal='center', vertical='center')
message = f"Count={count}"
cell_width = len(message) + 2
count_cell.value = message
std.column_dimensions['D'].width = cell_width

workbook.save('boohoo.xlsx')

# adjust based on location of the excel file
os.startfile('C:/Users/coope/Desktop/GitHub/T-Shirt-Web-Scraper/boohoo/boohoo.xlsx')

